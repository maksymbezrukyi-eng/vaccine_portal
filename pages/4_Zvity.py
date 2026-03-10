"""
Сторінка 4: Генерація звітів — Зведений файл (Level 0) та Level 1 для МОЗ
"""
import io
import streamlit as st
import pandas as pd

st.set_page_config(page_title="Звіти", page_icon="🏛️", layout="wide")
from datetime import datetime

if not st.session_state.get("authenticated"):
    st.error("⛔ Будь ласка, увійдіть через головну сторінку.")
    st.stop()

from core import (
    get_or_create_period, get_period_status, get_session,
    Organization, ReportFile,
)

st.title("🏛️ Генерація звітів")
st.markdown("Зведений файл по регіону та Level 1 файл для МОЗ/УЦКПХ.")
st.divider()

# ─── Сайдбар ──────────────────────────────────────────────────────────
with st.sidebar:
    st.header("📅 Звітний місяць")
    MONTHS = {1:"Січень",2:"Лютий",3:"Березень",4:"Квітень",5:"Травень",
              6:"Червень",7:"Липень",8:"Серпень",9:"Вересень",
              10:"Жовтень",11:"Листопад",12:"Грудень"}
    report_year  = st.selectbox("Рік:", range(2024, 2032), index=2)
    report_month = st.selectbox("Місяць:", list(MONTHS.keys()),
                                format_func=lambda x: MONTHS[x], index=0)
    period_label = f"{MONTHS[report_month]} {report_year}"

    st.divider()
    st.header("🏢 Реквізити ЦКПХ")
    ckkph_name   = st.text_input("Назва:", value="Житомирський обласний ЦКПХ")
    ckkph_edrpou = st.text_input("ЄДРПОУ:", value="")

    st.divider()
    if st.button("🚪 Вийти", use_container_width=True):
        st.session_state.clear()
        st.rerun()

period = get_or_create_period(report_year, report_month)
period_id = period.id

# Статус подання
session = get_session()
total_orgs = session.query(Organization).filter_by(is_active=True).count()
session.close()
status = get_period_status(period_id, total_orgs)

# Попередження
st.subheader(f"📅 {period_label}")
m1, m2, m3 = st.columns(3)
m1.metric("📥 Подали звіт", status["submitted"])
m2.metric("🟢 Придатних для зведення", status["ok"] + status["warning"])
m3.metric("🔴 З критичними помилками", status["error"])

good_count = status["ok"] + status["warning"]
if status["missing"] > 0:
    st.warning(f"⚠️ Не подали звіт: **{status['missing']}** ЗОЗ. Зведення буде неповним.")
if good_count == 0:
    st.error("❌ Немає файлів придатних для зведення.")
    st.stop()

st.divider()

# ─── Зведений файл (Level 0) ─────────────────────────────────────────
col_l0, col_l1 = st.columns(2)

with col_l0:
    st.subheader("📊 Зведений файл (Level 0)")
    st.markdown(f"""
    Агрегований регіональний файл у форматі Ф70.
    Включає **{good_count}** ЗОЗ з {status['submitted']} поданих.
    """)

    if st.button("⚙️ Сформувати зведений файл", type="primary",
                 use_container_width=True, key="btn_l0"):
        with st.spinner("⏳ Зведення виконується..."):
            try:
                # Агрегація з даних в БД
                from openpyxl import load_workbook, Workbook

                session = get_session()
                good_files = (
                    session.query(ReportFile)
                    .filter_by(period_id=period_id)
                    .filter(ReportFile.status.in_(["ok", "warning"]))
                    .all()
                )

                if not good_files:
                    st.error("Немає файлів для зведення.")
                    session.close()
                    st.stop()

                # Беремо шаблон з першого файлу
                template_bytes = good_files[0].file_bytes
                template_wb = load_workbook(io.BytesIO(template_bytes))

                def safe_n(v):
                    if v is None: return 0
                    try: return float(v)
                    except: return 0

                ws_exec_out = template_wb["Виконання"]
                ws_rem_out  = template_wb["Залишки"]
                ws_plan_out = template_wb["План"]
                ws_zvit_out = template_wb["Зведений звіт"]

                # Обнуляємо
                for row in range(8, 105):
                    ws_exec_out.cell(row, 5).value = 0
                for row in range(8, 11):
                    for col in [10,11,12,13,16,17,18]:
                        ws_exec_out.cell(row, col).value = 0
                for row in range(8, 14):
                    ws_exec_out.cell(row, 20).value = 0
                for row in range(11, 38):
                    for col in [2,3,5,6,7,8]:
                        ws_rem_out.cell(row, col).value = 0
                for row in range(11, 47):
                    ws_plan_out.cell(row, 6).value = 0
                for row in range(11, 120):
                    for col in [3,4,5]:
                        ws_zvit_out.cell(row, col).value = 0

                # Агрегуємо
                for rf in good_files:
                    wb = load_workbook(io.BytesIO(rf.file_bytes), data_only=True)
                    ws_e = wb["Виконання"]
                    ws_r = wb["Залишки"]
                    ws_p = wb["План"]
                    ws_z = wb["Зведений звіт"]

                    for row in range(8, 105):
                        ws_exec_out.cell(row, 5).value = (
                            ws_exec_out.cell(row, 5).value or 0) + safe_n(ws_e.cell(row, 5).value)
                    for row in range(8, 11):
                        for col in [10,11,12,13,16,17]:
                            ws_exec_out.cell(row, col).value = (
                                ws_exec_out.cell(row, col).value or 0) + safe_n(ws_e.cell(row, col).value)
                    for row in range(8, 14):
                        ws_exec_out.cell(row, 20).value = (
                            ws_exec_out.cell(row, 20).value or 0) + safe_n(ws_e.cell(row, 20).value)
                    for row in range(8, 11):
                        ws_exec_out.cell(row, 18).value = (
                            ws_exec_out.cell(row, 16).value or 0) + (ws_exec_out.cell(row, 17).value or 0)
                    for row in range(11, 38):
                        for col in [2,3,5,6,7,8]:
                            ws_rem_out.cell(row, col).value = (
                                ws_rem_out.cell(row, col).value or 0) + safe_n(ws_r.cell(row, col).value)
                    for row in range(11, 47):
                        ws_plan_out.cell(row, 6).value = (
                            ws_plan_out.cell(row, 6).value or 0) + safe_n(ws_p.cell(row, 6).value)
                    for row in range(11, 120):
                        for col in [3,4,5]:
                            ws_zvit_out.cell(row, col).value = (
                                ws_zvit_out.cell(row, col).value or 0) + safe_n(ws_z.cell(row, col).value)

                # Перераховуємо баланс залишків
                for row in range(11, 38):
                    if not ws_rem_out.cell(row, 1).value:
                        continue
                    b,c,f,g,h = [safe_n(ws_rem_out.cell(row, x).value) for x in [2,3,6,7,8]]
                    ws_rem_out.cell(row, 4).value = b+c+g+h-f

                # Перераховуємо %
                for row in range(11, 120):
                    c_val = ws_zvit_out.cell(row, 3).value
                    e_val = ws_zvit_out.cell(row, 5).value
                    ws_zvit_out.cell(row, 6).value = (
                        round(e_val / c_val * 100, 1)
                        if c_val and isinstance(c_val,(int,float)) and c_val > 0 and isinstance(e_val,(int,float))
                        else None
                    )

                # Реквізити
                rp = datetime(report_year, report_month, 1)
                template_wb["План"]["D8"].value = ckkph_name
                template_wb["План"]["E8"].value = ckkph_edrpou
                template_wb["Виконання"]["C4"].value = ckkph_name
                template_wb["Виконання"]["F4"].value = ckkph_edrpou
                template_wb["Виконання"]["F6"].value = rp
                template_wb["Залишки"]["A4"].value = ckkph_name
                template_wb["Залишки"]["D4"].value = ckkph_edrpou
                template_wb["Залишки"]["D6"].value = rp
                template_wb["Зведений звіт"]["A3"].value = ckkph_name
                template_wb["Зведений звіт"]["D3"].value = ckkph_edrpou
                template_wb["Зведений звіт"]["D5"].value = rp

                out = io.BytesIO()
                template_wb.save(out)
                out.seek(0)
                agg_bytes = out.getvalue()
                session.close()

                period_str = period_label.replace(" ", "_")
                st.session_state["agg_bytes"] = agg_bytes
                st.session_state["agg_name"]  = f"Зведений_Level0_{period_str}.xlsx"
                st.success(f"✅ Зведено {len(good_files)} файлів!")

            except Exception as e:
                st.error(f"❌ Помилка: {e}")
                import traceback
                st.code(traceback.format_exc())

    if "agg_bytes" in st.session_state:
        st.download_button(
            "⬇️ Завантажити зведений файл (Level 0)",
            data=st.session_state["agg_bytes"],
            file_name=st.session_state["agg_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary", key="dl_l0",
        )


# ─── Level 1 файл ────────────────────────────────────────────────────
with col_l1:
    st.subheader("🏛️ Level 1 файл для МОЗ")
    st.markdown(f"""
    Плоскі таблиці для передачі на національний рівень (МОЗ/ДУ УЦКПХ).
    Включає **7 аркушів** зі зведеними даними по всіх ЗОЗ.
    """)

    with st.expander("ℹ️ Структура Level 1 файлу"):
        st.markdown("""
| Аркуш | Вміст |
|---|---|
| `зведений` | Зведений звіт (вакцина/вік/план/місяць) по всіх ЗОЗ |
| `Аркуш2` | Деталізація виконання + народжуваність + протипокази + відмови |
| `Залишок` | Залишки вакцин усіх ЗОЗ |
| `Планування` | Плани вакцинації усіх ЗОЗ |
| `Аркуш3` | Народжуваність / КДП (1 рядок на ЗОЗ) |
| `Аркуш4` | Протипокази КДП (3 рядки на ЗОЗ) |
| `Аркуш5` | Відмови від щеплень (6 рядків на ЗОЗ) |
""")

    if st.button("⚙️ Сформувати Level 1 файл", type="primary",
                 use_container_width=True, key="btn_l1"):
        with st.spinner("⏳ Формування Level 1..."):
            try:
                from core.level1 import generate_level1
                lvl1_bytes = generate_level1(
                    period_id=period_id,
                    org_name=ckkph_name,
                    org_edrpou=ckkph_edrpou,
                )
                period_str = period_label.replace(" ", "_")
                st.session_state["lvl1_bytes"] = lvl1_bytes
                st.session_state["lvl1_name"]  = f"Level1_Житомирська_{period_str}.xlsx"
                st.success(f"✅ Level 1 файл сформовано!")
            except Exception as e:
                st.error(f"❌ Помилка: {e}")

    if "lvl1_bytes" in st.session_state:
        st.download_button(
            "⬇️ Завантажити Level 1",
            data=st.session_state["lvl1_bytes"],
            file_name=st.session_state["lvl1_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary", key="dl_l1",
        )

# ─── Перелік файлів включених у зведення ─────────────────────────────
st.divider()
st.subheader("📋 Файли включені в зведення")

session = get_session()
try:
    included = (
        session.query(ReportFile, Organization)
        .join(Organization, ReportFile.org_id == Organization.id)
        .filter(ReportFile.period_id == period_id)
        .filter(ReportFile.status.in_(["ok", "warning"]))
        .order_by(Organization.name)
        .all()
    )
    excluded = (
        session.query(ReportFile, Organization)
        .join(Organization, ReportFile.org_id == Organization.id)
        .filter(ReportFile.period_id == period_id)
        .filter(ReportFile.status == "error")
        .order_by(Organization.name)
        .all()
    )

    col_inc, col_exc = st.columns(2)

    with col_inc:
        st.markdown(f"**✅ Включено: {len(included)} файлів**")
        if included:
            rows = [{"Заклад": org.name, "Статус": "🟢 OK" if rf.status=="ok" else "🟡 Попередж."}
                    for rf, org in included]
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    with col_exc:
        st.markdown(f"**❌ Виключено: {len(excluded)} файлів (критичні помилки)**")
        if excluded:
            rows = [{"Заклад": org.name, "Помилок": rf.error_count}
                    for rf, org in excluded]
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        else:
            st.success("Виключених файлів немає — всі пройшли перевірку!")

finally:
    session.close()
