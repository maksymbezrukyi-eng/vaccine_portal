"""
level1.py — генерація Level1 файлу для МОЗ/УЦКПХ з даних БД
"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from .database import ReportFile, get_session


def _safe_num(val):
    if val is None:
        return 0
    try:
        return float(val)
    except Exception:
        return 0


def _hdr_style(ws, fill_hex="1F4E79"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(color="FFFFFF", bold=True, size=10)
    align = Alignment(wrap_text=True, vertical="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def generate_level1(period_id: int, org_name: str, org_edrpou: str) -> bytes:
    """
    Генерує Level1 файл із даних БД для вибраного звітного місяця.
    Повертає байти .xlsx файлу.
    """
    session = get_session()
    try:
        files = (
            session.query(ReportFile)
            .filter_by(period_id=period_id)
            .filter(ReportFile.status.in_(["ok", "warning"]))
            .all()
        )

        if not files:
            raise ValueError("Немає файлів для генерації Level1")

        # Визначаємо звітний місяць з першого файлу
        period = session.query(__import__('core.database', fromlist=['ReportPeriod']).ReportPeriod)\
            .filter_by(id=period_id).first()
        report_period = datetime(period.year, period.month, 1) if period else datetime.now()

        wb = Workbook()
        wb.remove(wb.active)

        H_ZVEDENY = [
            "Вакцина", "Вік",
            "Підлягали у звітному році",
            "Виконано за звітний місяць",
            "Назва закладу", "ЄДРПОУ", "Звітний період",
        ]
        H_ARKUSH2 = [
            "Назва закладу", "ЄДРПОУ", "Вакцина", "Вік",
            "Кількість щеплень", "Звітний період",
            "Народилось у пологовому", "ГепВ 1 доба",
            "Народилось за 7 місяців", "КДП3 своєчасно",
            "Антиген та вік (КДП)", "Кількість підлягаючих",
            "Тимчасові протипокази", "Постійні протипокази", "Всього протипоказів",
            "Нозологія відмов", "Кількість відмов",
        ]
        H_ZALISHOK = [
            "Вакцина", "Залишок на початок", "Отримано",
            "Залишок на кінець", "Виконано щеплень", "Використано",
            "Місц. бюджет", "Інші джерела",
            "Назва закладу", "ЄДРПОУ", "Звітний період",
        ]
        H_PLANUV = [
            "Назва закладу", "ЄДРПОУ", "Рік",
            "Нозологія", "Вік", "Кількість", "Примітка", "Вакцинація",
        ]
        H_ARKUSH3 = [
            "Назва закладу", "ЄДРПОУ", "Звітний період",
            "Народилось у пологовому", "ГепВ 1 доба",
            "Народилось за 7 місяців", "КДП3 своєчасно",
        ]
        H_ARKUSH4 = [
            "Назва закладу", "ЄДРПОУ", "Звітний період",
            "Антиген та вік", "Кількість підлягаючих",
            "Тимчасові", "Постійні", "Всього",
        ]
        H_ARKUSH5 = [
            "Нозологія відмов", "Кількість відмов",
            "Звітний період", "Назва закладу", "ЄДРПОУ",
        ]

        ws_zv  = wb.create_sheet("зведений")
        ws_a2  = wb.create_sheet("Аркуш2")
        ws_zal = wb.create_sheet("Залишок")
        ws_pl  = wb.create_sheet("Планування")
        ws_a3  = wb.create_sheet("Аркуш3")
        ws_a4  = wb.create_sheet("Аркуш4")
        ws_a5  = wb.create_sheet("Аркуш5")

        for ws, headers in [
            (ws_zv, H_ZVEDENY), (ws_a2, H_ARKUSH2), (ws_zal, H_ZALISHOK),
            (ws_pl, H_PLANUV), (ws_a3, H_ARKUSH3), (ws_a4, H_ARKUSH4), (ws_a5, H_ARKUSH5),
        ]:
            ws.append(headers)

        for rf in files:
            zoz_name = rf.organization.name if rf.organization else "—"
            zoz_edrpou = rf.organization.edrpou or "—" if rf.organization else "—"

            # Відкриваємо оригінальний файл
            if not rf.file_bytes:
                continue

            from openpyxl import load_workbook as lw
            orig_wb = lw(io.BytesIO(rf.file_bytes), data_only=True)
            ws_exec = orig_wb["Виконання"]
            ws_rem  = orig_wb["Залишки"]
            ws_zvit = orig_wb["Зведений звіт"]
            ws_plan = orig_wb["План"]

            # зведений
            for row in range(11, 120):
                vac = ws_zvit.cell(row, 1).value
                age = ws_zvit.cell(row, 2).value
                pl  = ws_zvit.cell(row, 3).value
                mon = ws_zvit.cell(row, 4).value
                if vac is None and mon is None:
                    continue
                ws_zv.append([vac, age, pl, mon, zoz_name, zoz_edrpou, report_period])

            # Аркуш2
            born8  = ws_exec.cell(8, 10).value
            hepv18 = ws_exec.cell(8, 11).value
            born78 = ws_exec.cell(8, 12).value
            kdp38  = ws_exec.cell(8, 13).value

            for row in range(8, 105):
                vac = ws_exec.cell(row, 3).value
                age = ws_exec.cell(row, 4).value
                cnt = ws_exec.cell(row, 5).value
                born  = born8  if row == 8 else None
                hepv1 = hepv18 if row == 8 else None
                born7 = born78 if row == 8 else None
                kdp3  = kdp38  if row == 8 else None
                kdp_lbl = ws_exec.cell(row, 14).value if row <= 10 else None
                kdp_cnt = ws_exec.cell(row, 15).value if row <= 10 else None
                temp_ci = ws_exec.cell(row, 16).value if row <= 10 else None
                perm_ci = ws_exec.cell(row, 17).value if row <= 10 else None
                all_ci  = ws_exec.cell(row, 18).value if row <= 10 else None
                ref_dis = ws_exec.cell(row, 19).value if row <= 13 else None
                ref_cnt = ws_exec.cell(row, 20).value if row <= 13 else None
                ws_a2.append([
                    zoz_name, zoz_edrpou, vac, age, cnt, report_period,
                    born, hepv1, born7, kdp3,
                    kdp_lbl, kdp_cnt, temp_ci, perm_ci, all_ci,
                    ref_dis, ref_cnt,
                ])

            # Аркуш3
            ws_a3.append([
                zoz_name, zoz_edrpou, report_period,
                born8, hepv18, born78, kdp38,
            ])

            # Аркуш4
            for row in range(8, 11):
                ws_a4.append([
                    zoz_name, zoz_edrpou, report_period,
                    ws_exec.cell(row, 14).value,
                    ws_exec.cell(row, 15).value,
                    ws_exec.cell(row, 16).value,
                    ws_exec.cell(row, 17).value,
                    ws_exec.cell(row, 18).value,
                ])

            # Аркуш5
            for row in range(8, 14):
                ws_a5.append([
                    ws_exec.cell(row, 19).value,
                    ws_exec.cell(row, 20).value,
                    report_period, zoz_name, zoz_edrpou,
                ])

            # Залишок
            for row in range(11, 38):
                vac = ws_rem.cell(row, 1).value
                if not vac:
                    continue
                ws_zal.append([
                    vac,
                    ws_rem.cell(row, 2).value, ws_rem.cell(row, 3).value,
                    ws_rem.cell(row, 4).value, ws_rem.cell(row, 5).value,
                    ws_rem.cell(row, 6).value, ws_rem.cell(row, 7).value,
                    ws_rem.cell(row, 8).value,
                    zoz_name, zoz_edrpou, report_period,
                ])

            # Планування
            for row in range(11, 47):
                noz = ws_plan.cell(row, 4).value
                if not noz:
                    continue
                ws_pl.append([
                    zoz_name, zoz_edrpou, report_period.year,
                    noz,
                    ws_plan.cell(row, 5).value,
                    ws_plan.cell(row, 6).value,
                    ws_plan.cell(row, 7).value,
                    ws_plan.cell(row, 8).value,
                ])

        # Стиль заголовків
        for ws in [ws_zv, ws_a2, ws_zal, ws_pl, ws_a3, ws_a4, ws_a5]:
            _hdr_style(ws)
            ws.freeze_panes = "A2"

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()

    finally:
        session.close()
