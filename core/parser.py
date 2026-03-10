"""
parser.py — парсинг та валідація Excel-файлів ЗОЗ (формат Ф70)
Адаптовано з vaccine-merger з виправленнями та розширеними перевірками.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Optional

from openpyxl import load_workbook


# ─── Утиліти ─────────────────────────────────────────────────────────

def safe_num(val: Any) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(",", ".").strip())
    except Exception:
        return 0.0


def safe_int(val: Any) -> int:
    return int(safe_num(val))


# ─── Результат парсингу ───────────────────────────────────────────────

class ParseResult:
    def __init__(self):
        self.ok = False
        self.name: str = "—"
        self.edrpou: str = "—"
        self.period: Optional[datetime] = None
        self.errors: list[str] = []
        self.warnings: list[str] = []
        self.execution_rows: list[dict] = []   # [{sheet_row, vaccine, age_group, planned, executed_month, ytd, pct}]
        self.stock_rows: list[dict] = []       # [{vaccine, stock_start, received, stock_end, executed, used, local_budget, other_sources}]
        self.refusal_rows: list[dict] = []     # [{disease, count}]
        self.birth_data: dict = {}             # {born_month, hepb_1day, born_7months, kdp3_timely, temp_ci, perm_ci}

    @property
    def status(self) -> str:
        if self.errors:
            return "error"
        if self.warnings:
            return "warning"
        return "ok"

    @property
    def status_emoji(self) -> str:
        return {"ok": "🟢 OK", "warning": "🟡 Попередження", "error": "🔴 Помилки"}[self.status]


# ─── Основна функція парсингу ─────────────────────────────────────────

def parse_file(file_bytes: bytes, filename: str = "") -> ParseResult:
    """
    Парсить та валідує один Excel-файл ЗОЗ.
    Повертає ParseResult з даними або з переліком помилок.
    """
    res = ParseResult()

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        res.errors.append(f"Не вдалось відкрити файл: {e}")
        return res

    # ── 1. Перевірка аркушів ─────────────────────────────────────────
    required_sheets = ["План", "Виконання", "Залишки", "Зведений звіт", "Аркуш1"]
    missing = [s for s in required_sheets if s not in wb.sheetnames]
    if missing:
        res.errors.append(f"Відсутні аркуші: {', '.join(missing)}")
        return res

    ws_plan = wb["План"]
    ws_exec = wb["Виконання"]
    ws_rem  = wb["Залишки"]
    ws_zvit = wb["Зведений звіт"]
    ws_ref  = wb["Аркуш1"]

    # ── 2. Назва закладу ─────────────────────────────────────────────
    # Первинне: Виконання!C4, резерв: Зведений!A3, потім Plan!D8
    raw_name = (
        ws_exec.cell(4, 3).value
        or ws_zvit.cell(3, 1).value
        or ws_plan["D8"].value
    )
    if raw_name:
        res.name = str(raw_name).strip()
    else:
        res.errors.append("Порожня назва закладу (Виконання!C4)")

    # ── 3. ЄДРПОУ ────────────────────────────────────────────────────
    raw_edrpou = (
        ws_exec.cell(4, 6).value
        or ws_zvit.cell(3, 4).value
        or ws_plan["E8"].value
    )
    if raw_edrpou:
        edrpou = str(raw_edrpou).strip().lstrip("'")
        res.edrpou = edrpou
        if not edrpou.isdigit() or len(edrpou) not in range(6, 11):
            res.warnings.append(f"ЄДРПОУ '{edrpou}' нестандартна довжина або не цифри")
    else:
        res.errors.append("Порожній ЄДРПОУ (Виконання!F4)")

    # ── 4. Звітний період ────────────────────────────────────────────
    raw_period = ws_exec.cell(6, 6).value
    if not raw_period:
        res.errors.append("Відсутній звітний період (Виконання!F6)")
    else:
        if isinstance(raw_period, datetime):
            res.period = raw_period
        else:
            res.errors.append(f"Некоректний формат звітного періоду: {raw_period}")

    # ── 5. Від'ємні значення у Виконанні ────────────────────────────
    for row in range(8, 105):
        val = ws_exec.cell(row, 5).value
        if isinstance(val, (int, float)) and val < 0:
            vac = ws_exec.cell(row, 3).value
            age = ws_exec.cell(row, 4).value
            res.errors.append(f"Від'ємна кількість щеплень: {vac}/{age} = {val}")

    # ── 6. Балансова формула залишків ────────────────────────────────
    for row in range(11, 38):
        vaccine = ws_rem.cell(row, 1).value
        if not vaccine:
            continue
        b = safe_num(ws_rem.cell(row, 2).value)
        c = safe_num(ws_rem.cell(row, 3).value)
        d = ws_rem.cell(row, 4).value
        f = safe_num(ws_rem.cell(row, 6).value)
        g = safe_num(ws_rem.cell(row, 7).value)
        h = safe_num(ws_rem.cell(row, 8).value)
        expected = b + c + g + h - f
        vname = str(vaccine).strip()
        if isinstance(d, (int, float)):
            if abs(d - expected) > 0.5:
                res.errors.append(
                    f"Залишки — помилка балансу «{vname}»: є {d}, має бути {round(expected, 1)}"
                )
            if d < 0:
                res.errors.append(f"Залишки — від'ємний залишок «{vname}»: {d}")
        # Використано > надійшло
        avail = b + c + g + h
        if f > avail + 0.5:
            res.errors.append(
                f"Залишки «{vname}»: використано ({int(f)}) > надійшло ({int(avail)})"
            )
        # Використано < виконано
        done = safe_num(ws_rem.cell(row, 5).value)
        if done > 0 and f < done - 0.5:
            res.warnings.append(
                f"Залишки «{vname}»: використано ({int(f)}) < виконано ({int(done)})"
            )

    # ── 7. Протипокази ───────────────────────────────────────────────
    for row in range(8, 11):
        p = safe_num(ws_exec.cell(row, 16).value)
        q = safe_num(ws_exec.cell(row, 17).value)
        r = ws_exec.cell(row, 18).value
        if isinstance(r, (int, float)) and abs(r - (p + q)) > 0.5:
            res.errors.append(
                f"Рядок {row}: протипокази ВСЬОГО ({r}) ≠ Тимчасові+Постійні ({p+q})"
            )

    # ── 8. КДП-3 своєчасність ────────────────────────────────────────
    for row in range(8, 11):
        l_val = safe_num(ws_exec.cell(row, 12).value)
        m_val = safe_num(ws_exec.cell(row, 13).value)
        if m_val > l_val > 0:
            res.errors.append(
                f"Рядок {row}: «КДП-3» ({int(m_val)}) > «Народилось за 7 міс» ({int(l_val)})"
            )

    # ── 9. Виконання ↔ Зведений звіт ────────────────────────────────
    EXEC_ZVIT_PAIRS = [
        (11, 15, "БЦЖ"), (23, 28, "Поліомієліт"), (35, 41, "Гепатит В"),
        (42, 49, "КПК"), (48, 56, "Hib"), (61, 70, "ВПЛ"),
        (99, 114, "АКДП"), (100, 115, "АаКДП"), (101, 116, "АДП"),
        (103, 118, "АДПм"), (104, 119, "АП"),
    ]
    for exec_row, zvit_row, label in EXEC_ZVIT_PAIRS:
        ev = safe_num(ws_exec.cell(exec_row, 7).value)
        zv = safe_num(ws_zvit.cell(zvit_row, 4).value)
        if abs(ev - zv) > 0.5 and (ev > 0 or zv > 0):
            res.errors.append(
                f"Розбіжність «{label}»: Виконання р{exec_row}={int(ev)} ≠ Зведений р{zvit_row}={int(zv)}"
            )

    # ── 10. % у Зведеному звіті ──────────────────────────────────────
    for row in range(11, 120):
        plan_v = ws_zvit.cell(row, 3).value
        ytd_v  = ws_zvit.cell(row, 5).value
        pct_v  = ws_zvit.cell(row, 6).value
        vac_v  = ws_zvit.cell(row, 1).value
        if (isinstance(plan_v, (int, float)) and plan_v > 0
                and isinstance(ytd_v, (int, float))
                and isinstance(pct_v, (int, float))):
            expected_pct = round(ytd_v / plan_v * 100, 2)
            if abs(pct_v - expected_pct) > 1.0:
                res.warnings.append(
                    f"Зведений р{row} «{str(vac_v or '').strip()}»: "
                    f"% у файлі={round(pct_v,1)}, розрахунковий={expected_pct}"
                )

    # ── 11. Когорт план-план ──────────────────────────────────────────
    COHORT_CHECKS = [
        (20, [34, 52, 81], "до 1 р (Поліо3=ГепВ3=Hib3=КДП3)"),
        (23, [85, 36, 53], "18 міс (Поліо4=КДП4=ГепВ4=Hib4)"),
        (26, [89], "6 р (Поліо5=АДП)"),
    ]
    for anchor_row, peer_rows, cohort_label in COHORT_CHECKS:
        anchor_val = safe_num(ws_zvit.cell(anchor_row, 3).value)
        if anchor_val <= 0:
            continue
        for peer_row in peer_rows:
            peer_val = safe_num(ws_zvit.cell(peer_row, 3).value)
            if peer_val <= 0:
                continue
            if abs(anchor_val - peer_val) / anchor_val > 0.01:
                res.warnings.append(
                    f"Розбіжність планів когорту «{cohort_label}»: "
                    f"р{anchor_row}={int(anchor_val)} ≠ р{peer_row}={int(peer_val)}"
                )

    # ═══════════════════════════════════════════════════════════════════
    # ЗБІР ДАНИХ ДЛЯ БД
    # ═══════════════════════════════════════════════════════════════════

    # ── Дані Зведений звіт ───────────────────────────────────────────
    for row in range(11, 120):
        vaccine  = ws_zvit.cell(row, 1).value
        age      = ws_zvit.cell(row, 2).value
        planned  = ws_zvit.cell(row, 3).value
        month_ex = ws_zvit.cell(row, 4).value
        ytd      = ws_zvit.cell(row, 5).value
        pct      = ws_zvit.cell(row, 6).value
        if vaccine is None:
            continue
        res.execution_rows.append({
            "sheet_row":      row,
            "vaccine":        str(vaccine).strip(),
            "age_group":      str(age or "").strip(),
            "planned":        safe_num(planned) if isinstance(planned, (int, float)) else None,
            "executed_month": safe_num(month_ex) if isinstance(month_ex, (int, float)) else None,
            "ytd":            safe_num(ytd) if isinstance(ytd, (int, float)) else None,
            "pct":            safe_num(pct) if isinstance(pct, (int, float)) else None,
        })

    # ── Залишки ──────────────────────────────────────────────────────
    for row in range(11, 38):
        vaccine = ws_rem.cell(row, 1).value
        if not vaccine:
            continue
        res.stock_rows.append({
            "vaccine":      str(vaccine).strip(),
            "stock_start":  safe_num(ws_rem.cell(row, 2).value),
            "received":     safe_num(ws_rem.cell(row, 3).value),
            "stock_end":    safe_num(ws_rem.cell(row, 4).value),
            "executed":     safe_num(ws_rem.cell(row, 5).value),
            "used":         safe_num(ws_rem.cell(row, 6).value),
            "local_budget": safe_num(ws_rem.cell(row, 7).value),
            "other_sources":safe_num(ws_rem.cell(row, 8).value),
        })

    # ── Відмови ──────────────────────────────────────────────────────
    REFUSAL_MAP = {
        8: "Туберкульоз", 9: "Поліомієліт", 10: "Гепатит В",
        11: "Кашлюк, дифтерія, правець",
        12: "Гемофільна інфекція",
        13: "Кір, паротит, краснуха",
    }
    for row, disease in REFUSAL_MAP.items():
        count = safe_int(ws_exec.cell(row, 20).value)
        res.refusal_rows.append({"disease": disease, "count": count})

    # ── Народжуваність / протипокази ─────────────────────────────────
    res.birth_data = {
        "born_month":   safe_int(ws_exec.cell(8, 10).value),
        "hepb_1day":    safe_int(ws_exec.cell(8, 11).value),
        "born_7months": safe_int(ws_exec.cell(8, 12).value),
        "kdp3_timely":  safe_int(ws_exec.cell(8, 13).value),
        "temp_ci": sum(safe_int(ws_exec.cell(r, 16).value) for r in range(8, 11)),
        "perm_ci": sum(safe_int(ws_exec.cell(r, 17).value) for r in range(8, 11)),
    }

    res.ok = True
    return res
